import streamlit as st
import pdfplumber
import re
import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)

# ── Colores Quálitas ──────────────────────────────────────────────────────────
TEAL   = "FF006B6B"   # encabezado oscuro
LTEAL  = "FFE0F4F4"   # fila par
WHITE  = "FFFFFFFF"
RED    = "FFC00000"   # acento
GRAY   = "FF595959"

COLS = [
    "Fecha", "Hora", "N° Reporte", "N° Póliza",
    "Nombre", "Teléfono", "E-mail",
    "Marca", "Tipo", "Modelo (Año)", "Color",
    "Descripción de Daños",
]

# ── Helpers de extracción ─────────────────────────────────────────────────────

def clean(txt):
    return re.sub(r"\s+", " ", txt or "").strip()


def first_match(pattern, text, group=1, flags=re.IGNORECASE):
    m = re.search(pattern, text, flags)
    return clean(m.group(group)) if m else ""


# ── Formato 1: Orden de Admisión Automóviles ──────────────────────────────────

def parse_automoviles(text):
    # Fecha
    fecha = first_match(
        r"FECHA\s*/\s*[^/\n]*?\s*(\d{2}/\d{2}/\d{4})", text
    ) or first_match(r"(\d{2}/\d{2}/\d{4})", text)

    # Hora
    hora = first_match(r"(\d{1,2}:\d{2})\s*HRS", text)

    # Reporte
    reporte = first_match(r"N[°º]\.\s*REPORTE\s+(\d+)", text)

    # Póliza  (primera secuencia larga de dígitos en la línea de póliza)
    poliza = first_match(
        r"N[°º]\s*DE\s*P[OÓ]LIZA[^/\n]*/[^/\n]*/\s*(\d+)", text
    ) or first_match(r"(\d{10,})", text)

    # Nombre asegurado
    nombre = first_match(
        r"(?:ASEGURADO|TERCERO\s*Q)\s*/?\s*\n([A-ZÁÉÍÓÚÑ ]{5,})", text
    ) or first_match(
        r"NOMBRE\s*O\s*RAZ[OÓ]N\s*SOCIAL\s*DEL\s*CLIENTE\s*/[^\n]*\n([A-ZÁÉÍÓÚÑ ]{5,})", text
    )

    # Teléfono (primer número de 10 dígitos con posible espacio)
    tel = first_match(r"TEL[EÉ]FONO\s*/?\s*\n?([\d\s]{10,14})", text)
    tel = re.sub(r"\s", "", tel)

    # Email
    email = first_match(r"E-?MAIL\s*/?\s*\n?([\w.\-+]+@[\w.\-]+)", text)

    # Vehículo
    marca = first_match(r"MARCA\s*/[^\n]*\n(\w+)", text)
    tipo  = first_match(r"TIPO\s*/[^\n]*\n([A-Z0-9 ]+)", text)
    modelo= first_match(r"MODELO\s*\([AÑO]+\)\s*/[^\n]*\n(\d{4})", text)
    color = first_match(r"COLOR\s*\n([A-ZÁÉÍÓÚÑ]+)", text)

    # Descripción de daños
    desc  = first_match(
        r"DESCRIPCI[OÓ]N\s*DE\s*DA[ÑN]OS\s*A\s*REPARAR[^\n]*\n(.+?)(?:\n[A-Z]{2,}|\Z)",
        text, flags=re.IGNORECASE | re.DOTALL
    )

    return {
        "Fecha": fecha, "Hora": hora, "N° Reporte": reporte,
        "N° Póliza": poliza, "Nombre": nombre,
        "Teléfono": tel, "E-mail": email,
        "Marca": marca, "Tipo": tipo, "Modelo (Año)": modelo,
        "Color": color, "Descripción de Daños": desc,
    }


# ── Formato 2: Orden de Admisión Ajuste Express ───────────────────────────────

def parse_express(text):
    # Fecha  (formato YYYY-MM-DD o DD/MM/YYYY)
    fecha = first_match(r"FECHA\s+(\d{4}-\d{2}-\d{2})", text)
    if fecha:
        # Convertir a DD/MM/YYYY
        parts = fecha.split("-")
        fecha = f"{parts[2]}/{parts[1]}/{parts[0]}"

    # Hora → no existe en Express
    hora = ""

    # Reporte / Siniestro
    reporte = first_match(r"N[°º]\.\s*REPORTE\s+(\d+)", text)

    # Póliza → no presente en Express
    poliza = ""

    # Nombre asegurado (firma del conductor)
    nombre = first_match(
        r"FIRMA\s*DEL\s*CONDUCTOR\s*ASEGURADO?\s*\n([A-ZÁÉÍÓÚÑ ]{5,})", text
    ) or first_match(r"ASEGURADO\s*\n([A-ZÁÉÍÓÚÑ ]{5,})", text)

    # Teléfono / Email → no presentes
    tel   = ""
    email = ""

    # Vehículo
    marca  = first_match(r"^(MAZDA|FORD|CHEVROLET|KIA|NISSAN|TOYOTA|VOLKSWAGEN|HONDA|HYUNDAI|CHRYSLER|DODGE|JEEP|RAM|SEAT|RENAULT|MITSUBISHI|SUZUKI|SUBARU|VOLVO|BMW|MERCEDES|AUDI|PEUGEOT|FIAT|ACURA|INFINITI|LEXUS|CADILLAC|BUICK|GMC|LINCOLN)\b", text, flags=re.MULTILINE)
    tipo   = ""
    modelo = ""
    color  = ""

    # Buscar bloque de tabla vehículo
    veh_m = re.search(
        r"(MAZDA|FORD|CHEVROLET|KIA|NISSAN|TOYOTA|VOLKSWAGEN|HONDA|HYUNDAI|CHRYSLER|DODGE|JEEP|RAM|SEAT|RENAULT|MITSUBISHI|SUZUKI|SUBARU|VOLVO|BMW|MERCEDES|AUDI|PEUGEOT|FIAT|ACURA|INFINITI|LEXUS|CADILLAC|BUICK|GMC|LINCOLN)\s+([A-Z0-9 ]+?)\s+(\d{4})\s+\d",
        text, re.IGNORECASE
    )
    if veh_m:
        marca  = veh_m.group(1).strip()
        tipo   = veh_m.group(2).strip()
        modelo = veh_m.group(3).strip()

    color_m = re.search(r"(NEGRO|BLANCO|ROJO|AZUL|GRIS|PLATA|PLATEADO|VERDE|AMARILLO|NARANJA|CAFE|CAFÉ|MORADO|BEIGE|VINO|DORADO|ROSA)\b", text, re.IGNORECASE)
    if color_m:
        color = color_m.group(1).upper()

    # Descripción de daños
    desc = first_match(
        r"Descripci[oó]n\s*de\s*da[ñn]os\s*\n(.+?)(?:\nDa[ñn]os\s*Preexistentes|\Z)",
        text, flags=re.IGNORECASE | re.DOTALL
    )

    return {
        "Fecha": fecha, "Hora": hora, "N° Reporte": reporte,
        "N° Póliza": poliza, "Nombre": nombre,
        "Teléfono": tel, "E-mail": email,
        "Marca": marca, "Tipo": tipo, "Modelo (Año)": modelo,
        "Color": color, "Descripción de Daños": desc,
    }


# ── Detección de formato y parseo ─────────────────────────────────────────────

def extract_from_pdf(uploaded_file):
    with pdfplumber.open(uploaded_file) as pdf:
        text = "\n".join(p.extract_text() or "" for p in pdf.pages)

    is_express = bool(re.search(r"AJUSTE\s*EXPRESS", text, re.IGNORECASE))
    data = parse_express(text) if is_express else parse_automoviles(text)
    data["_tipo"] = "Express" if is_express else "Automóviles"
    data["_texto"] = text
    return data


# ── Generación de Excel ───────────────────────────────────────────────────────

def make_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Admisiones Quálitas"

    # Estilos
    hdr_fill  = PatternFill("solid", fgColor=TEAL)
    hdr_font  = Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
    row_even  = PatternFill("solid", fgColor=LTEAL)
    row_odd   = PatternFill("solid", fgColor=WHITE)
    border    = Border(
        left=Side(style="thin", color="FFCCCCCC"),
        right=Side(style="thin", color="FFCCCCCC"),
        top=Side(style="thin", color="FFCCCCCC"),
        bottom=Side(style="thin", color="FFCCCCCC"),
    )
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Título
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "Órdenes de Admisión — Quálitas"
    title_cell.font  = Font(bold=True, color="FFFFFFFF", name="Arial", size=13)
    title_cell.fill  = PatternFill("solid", fgColor=RED)
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # Encabezados
    for col_idx, col_name in enumerate(COLS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[2].height = 22

    # Datos
    for r_idx, row in enumerate(rows, start=3):
        fill = row_even if r_idx % 2 == 0 else row_odd
        for c_idx, col in enumerate(COLS, start=1):
            val  = row.get(col, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill      = fill
            cell.border    = border
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = left if c_idx in (5, 7, 12) else center

    # Anchos de columna
    from openpyxl.utils import get_column_letter
    widths = [12, 10, 15, 16, 28, 16, 28, 12, 24, 14, 12, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── UI Streamlit ──────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Admisiones Quálitas",
    page_icon="🚗",
    layout="wide",
)

# CSS personalizado
st.markdown("""
<style>
    .main { background-color: #F5FAFA; }
    h1 { color: #006B6B; }
    .stButton > button {
        background-color: #006B6B;
        color: white;
        border-radius: 6px;
        font-weight: bold;
        padding: 0.4rem 1.2rem;
    }
    .stButton > button:hover { background-color: #004F4F; }
    .block-container { padding-top: 2rem; }
    .stDataFrame { border-radius: 8px; }
</style>
""", unsafe_allow_html=True)

# Encabezado
col_logo, col_title = st.columns([1, 6])
with col_logo:
    st.markdown("## 🚗")
with col_title:
    st.title("Extractor de Órdenes de Admisión — Quálitas")
    st.caption("Sube una o más órdenes en PDF y descarga los datos en Excel")

st.divider()

uploaded = st.file_uploader(
    "Arrastra aquí tus PDFs de Quálitas",
    type=["pdf"],
    accept_multiple_files=True,
    help="Se aceptan tanto 'Orden de Admisión Automóviles' como 'Ajuste Express'",
)

if uploaded:
    rows   = []
    errors = []

    with st.spinner("Procesando PDFs…"):
        for f in uploaded:
            try:
                data = extract_from_pdf(f)
                tipo = data.pop("_tipo", "")
                data.pop("_texto", None)
                # Renombrar clave interna al nombre de columna
                row = {col: data.get(col, "") for col in COLS}
                row["_tipo"] = tipo
                rows.append(row)
            except Exception as e:
                errors.append(f"**{f.name}**: {e}")

    if errors:
        for err in errors:
            st.error(err)

    if rows:
        st.success(f"✅ {len(rows)} PDF(s) procesados correctamente")

        # Tabla previa
        import pandas as pd
        df = pd.DataFrame(rows)

        # Indicador de tipo
        tipo_col = df.pop("_tipo")
        st.markdown("#### Vista previa")
        st.dataframe(
            df,
            use_container_width=True,
            height=min(200 + len(rows) * 38, 500),
        )

        # Detalle expandible por orden
        with st.expander("🔍 Ver detalle por orden"):
            for i, row in enumerate(rows):
                t = tipo_col.iloc[i]
                badge = "🔵 Automóviles" if t == "Automóviles" else "🟣 Ajuste Express"
                st.markdown(f"**{badge} — Reporte {row.get('N° Reporte', i+1)}**")
                cols = st.columns(3)
                items = [(k, v) for k, v in row.items() if k != "_tipo"]
                for j, (k, v) in enumerate(items):
                    cols[j % 3].markdown(f"- **{k}:** {v or '—'}")
                st.divider()

        # Botón de descarga Excel
        excel_buf = make_excel(rows)
        st.download_button(
            label="⬇️  Descargar Excel",
            data=excel_buf,
            file_name="admisiones_qualitas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

else:
    st.info("👆 Sube uno o más PDFs de Quálitas para comenzar.")
    st.markdown("""
    **Formatos compatibles:**
    - 📄 Orden de Admisión Automóviles
    - 📄 Orden de Admisión Ajuste Express

    **Campos que se extraen:**
    `Fecha · Hora · N° Reporte · N° Póliza · Nombre · Teléfono · E-mail · Marca · Tipo · Modelo · Color · Descripción de Daños`
    """)

